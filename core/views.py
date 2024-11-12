import base64
from typing import List

import openpyxl
import requests
from django.conf import settings
from django.contrib.auth import authenticate
from django.db.models import F, FloatField, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (Acrescimo, AuthenticationToken, Caixa, Categoria,
                     Delivery, Despesa, Funcionario, Mesa, MetodoPagamento,
                     Notificacao, Pedido, Pessoa, Prato, PratoPedido,
                     Restaurante, User)
from .serializers import (AcrescimoSerializer, CaixaSerializer,
                          CategoriaSerializer, DeliverySerializer,
                          DespesaSerializer, MesaEPedidosSerializer,
                          MesaSerializer, MetodoPagamentoSerializer,
                          NotificacaoSerializer, PedidoSerializer,
                          PessoaSerializer, PratoPedidoSerializer,
                          PratoSerializer, PratoSimpleSerializer,
                          RestauranteSerializer)

AUTH_STATE = {}


# class PedidoWhatsApp(APIView):
#     @staticmethod
#     def post(self, request): ...


class Item:
    def __init__(
        self, codigo: str, descricao: str, unidade: str, quantidade: float, valor: float
    ):
        self.codigo = codigo
        self.descricao = descricao
        self.unidade = unidade
        self.quantidade = quantidade
        self.valor = valor


class BlingNFCeService:
    @staticmethod
    def enviar_nfce(itens: List[Item]):
        access_token = BlingTokenService.get_access_token()
        if not access_token:
            return {"error": "Não foi possível autenticar com a API Bling"}

        # Estrutura de dados para NFC-e
        nfce_data = {
            "tipo": 1,  # Tipo de nota fiscal
            "dataOperacao": "2023-01-12 09:52:12",  # Data da operação (você pode ajustar conforme necessário)
            "contato": {
                "nome": "Consumidor final",
                "tipoPessoa": "F",
            },
            "naturezaOperacao": {
                "id": 12345678  # ID da operação, altere conforme sua configuração no Bling
            },
            "itens": [
                {
                    "codigo": item.codigo,
                    "descricao": item.descricao,
                    "unidade": item.unidade,
                    "quantidade": item.quantidade,
                    "valor": item.valor,
                    "tipo": "P",  # Tipo de produto
                    "origem": 0,
                }
                for item in itens
            ],
            "finalidade": 1,  # Finalidade da emissão
            "observacoes": "Nota fiscal de venda ao consumidor final.",
        }

        url = "https://api.bling.com.br/Api/v3/nfce"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}",
        }

        # Envia a NFC-e como JSON
        response = requests.post(url, json=nfce_data, headers=headers)

        # Verifica se a resposta foi bem-sucedida e retorna a resposta
        if response.status_code == 200:
            return response.json()
        else:
            return {"error": response.text}


class aaa(APIView):
    def get(self, request):
        itens = [
            Item("001", "Produto Teste 1", "UN", 2, 10.0),
            Item("002", "Produto Teste 2", "UN", 1, 20.0),
        ]
        resposta = BlingNFCeService.enviar_nfce(itens)

        return Response(resposta)


class BlingTokenService:
    @staticmethod
    def get_access_token():
        # Tenta obter um token válido do banco de dados
        token = AuthenticationToken.get_valid_access_token()
        if token:
            return token

        # Se não houver token válido, tenta usar o refresh_token para obter um novo access_token
        refresh_token = BlingTokenService.get_refresh_token()
        if refresh_token:
            return BlingTokenService.refresh_access_token(refresh_token)

        return None

    @staticmethod
    def store_tokens(access_token, refresh_token, expires_in):
        # Salva ou atualiza o token de autenticação no banco de dados
        AuthenticationToken.objects.create(
            access_token=access_token, refresh_token=refresh_token, expires_in=expires_in
        )

    @staticmethod
    def get_refresh_token():
        # Tenta obter o refresh_token mais recente do banco de dados
        token = (
            AuthenticationToken.objects.filter(expires_at__gt=timezone.now())
            .order_by("-expires_at")
            .first()
        )
        return token.refresh_token if token else None

    @staticmethod
    def refresh_access_token(refresh_token):
        token_url = "https://bling.com.br/Api/v3/oauth/token"
        auth_string = f"{settings.BLING_CLIENT_ID}:{settings.BLING_CLIENT_SECRET}"
        auth_base64 = base64.b64encode(auth_string.encode()).decode()

        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "1.0",
            "Authorization": f"Basic {auth_base64}",
        }
        data = {"grant_type": "refresh_token", "refresh_token": refresh_token}

        response = requests.post(token_url, data=data, headers=headers)
        if response.status_code == 200:
            tokens = response.json()
            # Salva o novo access_token e refresh_token no banco de dados
            BlingTokenService.store_tokens(
                tokens["access_token"], tokens["refresh_token"], tokens["expires_in"]
            )
            return tokens["access_token"]
        return None


class BlingOAuthCallbackView(APIView):
    def get(self, request):
        code = request.GET.get("code")

        if not code:
            return Response(
                {"error": "Code não fornecido"}, status=status.HTTP_400_BAD_REQUEST
            )

        token_url = "https://bling.com.br/Api/v3/oauth/token"
        auth_string = f"{settings.BLING_CLIENT_ID}:{settings.BLING_CLIENT_SECRET}"
        auth_base64 = base64.b64encode(auth_string.encode()).decode()

        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "1.0",
            "Authorization": f"Basic {auth_base64}",
        }
        data = {"grant_type": "authorization_code", "code": code}

        response = requests.post(token_url, data=data, headers=headers)

        if response.status_code == 200:
            tokens = response.json()
            BlingTokenService.store_tokens(
                tokens["access_token"], tokens["refresh_token"], tokens["expires_in"]
            )
            return Response(tokens, status=status.HTTP_200_OK)
        elif response.status_code == 429:
            return Response(status=status.HTTP_429_TOO_MANY_REQUESTS)
        return Response(response.json(), status=status.HTTP_400_BAD_REQUEST)


class DeliveryManagementViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.filter(is_delivery=True)
    serializer_class = PedidoSerializer

    def get_queryset(self):
        restaurante_id = self.request.query_params.get("restaurante_id")
        if restaurante_id:
            return self.queryset.filter(caixa__restaurante_id=restaurante_id)
        return self.queryset

    def list(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(is_delivery=True)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"])
    def cancelar(self, request, pk=None):
        pedido = self.get_object()
        pedido.status = "Cancelado"
        pedido.save()
        return Response({"status": "Pedido cancelado"})

    @action(detail=False, methods=["get"])
    def clientes(self, request):
        restaurante_id = request.query_params.get("restaurante_id")
        clientes = Pessoa.objects.filter(
            pedido__is_delivery=True, restaurante_id=restaurante_id
        ).distinct()
        serializer = PessoaSerializer(clientes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def produtos(self, request):
        restaurante_id = request.query_params.get("restaurante_id")
        produtos = Prato.objects.filter(restaurante_id=restaurante_id)
        serializer = PratoSerializer(produtos, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def detalhes(self, request, pk=None):
        pedido = self.get_object()
        serializer = self.get_serializer(pedido)
        return Response(serializer.data)


class PedidoAtualMesaView(APIView):
    def get(self, request, restaurante_id, mesa_id):
        # Busca a mesa e o pedido atual associado à mesa e ao restaurante
        mesa = get_object_or_404(Mesa, id=mesa_id, restaurante_id=restaurante_id)
        pedido = Pedido.objects.filter(
            mesa=mesa, valor_pago__isnull=True
        ).last()  # Busca o último pedido não pago

        if not pedido:
            return Response(
                {"error": "Nenhum pedido aberto para esta mesa."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Organiza os detalhes dos itens do pedido
        itens = []
        subtotal = 0
        for prato_pedido in pedido.pratos.all():
            item_total = prato_pedido.prato.valor * prato_pedido.quantidade
            subtotal += item_total
            itens.append(
                {
                    "nome": f"{prato_pedido.prato.nome} {prato_pedido.quantidade}x",
                    "valor": f"R${item_total:.2f}",
                }
            )

        # Resumo do pedido
        total = subtotal

        data = {
            "mesa": f"Mesa {mesa.numero_mesa}",
            "itens": itens,
            "pedido_id": pedido.id,
            "subtotal": f"R${subtotal:.2f}",
            "total": f"R${total:.2f}",
            "status_pedido": pedido.status,
        }

        return Response(data, status=status.HTTP_200_OK)


class ExportarTransacoesExcelView(APIView):
    def get(self, request, restaurante_id, caixa_id):
        # Verifica se o restaurante existe
        try:
            restaurante = Restaurante.objects.get(id=restaurante_id)
        except Restaurante.DoesNotExist:
            return Response(
                {"error": "Restaurante não encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Verifica se o caixa existe para o restaurante e está aberto
        try:
            caixa = Caixa.objects.get(id=caixa_id, restaurante=restaurante)
        except Caixa.DoesNotExist:
            return Response(
                {"error": "Caixa não encontrado para este restaurante."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Criando o arquivo Excel
        wb = openpyxl.Workbook()

        # Planilha de Pedidos
        pedidos_ws = wb.active
        pedidos_ws.title = "Pedidos"
        pedidos_ws.append(
            [
                "ID",
                "Pessoa",
                "Valor Pago",
                "Desconto",
                "Troco",
                "Subtotal",
                "Status",
                "Horário",
            ]
        )
        pedidos = Pedido.objects.filter(caixa=caixa)
        subtotal = 0
        for pedido in pedidos:
            if pedido.status == "Em Confirmacao":
                continue
            for ordem in pedido.pratos.all():
                subtotal += ordem.prato.valor

            pedidos_ws.append(
                [
                    str(pedido.id),
                    pedido.pessoa.nome if pedido.pessoa else "Anônimo",
                    pedido.valor_pago or 0,
                    pedido.desconto or 0,
                    pedido.troco or 0,
                    pedido.subtotal or pedido.valor_pago or subtotal,
                    pedido.status,
                    pedido.horario_pedido.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        # Ajusta a largura das colunas de pedidos
        for col in pedidos_ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            pedidos_ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max_length
            )

        # Planilha de Despesas
        despesas_ws = wb.create_sheet(title="Despesas")
        despesas_ws.append(["ID", "Descrição", "Valor", "Categoria", "Data de Criação"])
        despesas = Despesa.objects.filter(caixa=caixa)
        for despesa in despesas:
            despesas_ws.append(
                [
                    despesa.id,
                    despesa.descricao,
                    despesa.valor,
                    despesa.categoria,
                    despesa.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        # Ajusta a largura das colunas de despesas
        for col in despesas_ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            despesas_ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max_length
            )

        # Planilha de Acréscimos
        acrescimos_ws = wb.create_sheet(title="Acréscimos")
        acrescimos_ws.append(["ID", "Descrição", "Valor", "Categoria", "Data de Criação"])
        acrescimos = Acrescimo.objects.filter(caixa=caixa)
        for acrescimo in acrescimos:
            acrescimos_ws.append(
                [
                    acrescimo.id,
                    acrescimo.descricao,
                    acrescimo.valor,
                    acrescimo.categoria,
                    acrescimo.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        # Ajusta a largura das colunas de acréscimos
        for col in acrescimos_ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            acrescimos_ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max_length
            )

        # Salvando o arquivo em memória para envio
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="transacoes_restaurante_{restaurante_id}_caixa_{caixa_id}.xlsx"'
        )
        wb.save(response)
        return response


class ResumoVendasView(APIView):
    def get(self, request, restaurante_id, caixa_id):
        # Tenta buscar o restaurante com base no ID fornecido
        try:
            restaurante = Restaurante.objects.get(id=restaurante_id)
        except Restaurante.DoesNotExist:
            return Response(
                {"error": "Restaurante não encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Tenta buscar o caixa com base no restaurante e no caixa_id
        try:
            caixa = Caixa.objects.get(
                id=caixa_id, restaurante=restaurante, fechado_em__isnull=True
            )
        except Caixa.DoesNotExist:
            return Response(
                {"error": "Caixa aberto não encontrado para este restaurante."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Calcula o total faturado (considerando o valor dos métodos de pagamento)
        total_faturado = MetodoPagamento.objects.filter(pedidos__caixa=caixa).aggregate(
            total=Coalesce(
                Sum(F("valor"), output_field=FloatField()),
                Value(0, output_field=FloatField()),
            )
        )["total"]

        # Calcula o valor recebido por método de pagamento
        pagamentos_por_metodo = (
            MetodoPagamento.objects.filter(pedidos__caixa=caixa)
            .values("metodo")
            .annotate(
                total=Coalesce(
                    Sum(F("valor"), output_field=FloatField()),
                    Value(0, output_field=FloatField()),
                )
            )
            .order_by("-total")
        )

        # Formata os métodos de pagamento com valores em BRL
        pagamentos_formatados = {
            pagamento["metodo"].replace("_", " ").title(): "R$" + pagamento["total"]
            for pagamento in pagamentos_por_metodo
        }

        # Calcula o total de despesas
        total_despesas = Despesa.objects.filter(caixa=caixa).aggregate(
            total=Coalesce(
                Sum(F("valor"), output_field=FloatField()),
                Value(0, output_field=FloatField()),
            )
        )["total"]

        # Ranking das categorias de despesas
        ranking_despesas = (
            Despesa.objects.filter(caixa=caixa)
            .values("categoria")
            .annotate(
                total=Coalesce(
                    Sum(F("valor"), output_field=FloatField()),
                    Value(0, output_field=FloatField()),
                )
            )
            .order_by("-total")
        )
        ranking_despesas_formatado = {
            despesa["categoria"].title(): "R$" + despesa["total"]
            for despesa in ranking_despesas
        }

        # Calcula o total de acréscimos
        total_acrescimos = Acrescimo.objects.filter(caixa=caixa).aggregate(
            total=Coalesce(
                Sum(F("valor"), output_field=FloatField()),
                Value(0, output_field=FloatField()),
            )
        )["total"]

        # Ranking das categorias de acréscimos
        ranking_acrescimos = (
            Acrescimo.objects.filter(caixa=caixa)
            .values("categoria")
            .annotate(
                total=Coalesce(
                    Sum(F("valor"), output_field=FloatField()),
                    Value(0, output_field=FloatField()),
                )
            )
            .order_by("-total")
        )
        ranking_acrescimos_formatado = {
            acrescimo["categoria"].title(): "R$" + acrescimo["total"]
            for acrescimo in ranking_acrescimos
        }

        # Calcula o total em caixa
        saldo_inicial = caixa.saldo_inicial or 0
        total_em_caixa = (
            saldo_inicial + total_faturado + total_acrescimos - total_despesas
        )

        # Monta o dicionário de resposta com valores formatados
        response_data = {
            "total_faturado": "R$" + total_faturado,
            "pagamentos_por_metodo": pagamentos_formatados,
            "total_despesas": "R$" + total_despesas,
            "ranking_despesas": ranking_despesas_formatado,
            "total_acrescimos": "R$" + total_acrescimos,
            "ranking_acrescimos": ranking_acrescimos_formatado,
            "total_em_caixa": "R$" + total_em_caixa,
        }

        return Response(response_data, status=status.HTTP_200_OK)


class CaixaAbertoView(APIView):
    def get(self, request):
        # Obtém o parâmetro 'restaurante_id' da requisição
        restaurante_id = request.query_params.get("restaurante_id")

        # Verifica se o parâmetro 'restaurante_id' foi fornecido
        if not restaurante_id:
            return Response(
                {"error": "Parâmetro 'restaurante_id' é obrigatório."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Tenta buscar o restaurante com base no ID fornecido
        try:
            restaurante = Restaurante.objects.get(id=restaurante_id)
        except Restaurante.DoesNotExist:
            return Response(
                {"error": "Restaurante não encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Tenta buscar o caixa aberto para o restaurante
        caixa_aberto = Caixa.objects.filter(
            restaurante=restaurante, fechado_em__isnull=True
        ).first()

        if caixa_aberto:
            # Retorna o ID do caixa aberto
            return Response(
                {"caixa_aberto_id": caixa_aberto.id}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"error": "Não há caixa aberto para este restaurante."},
                status=status.HTTP_404_NOT_FOUND,
            )


class LogoutView(APIView):
    def post(self, request):
        try:
            # Pega o token de refresh do corpo da requisição
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                return Response(
                    {"error": "Token de refresh é obrigatório."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Invalida o token
            token = RefreshToken(refresh_token)
            token.blacklist()

            return Response(
                {"success": "Logout realizado com sucesso."},
                status=status.HTTP_205_RESET_CONTENT,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class UsuarioDetalhesView(APIView):
    def get(self, request):
        # Obtém o parâmetro 'id' da requisição
        user_id = request.query_params.get("id")

        # Verifica se o parâmetro 'id' foi fornecido
        if not user_id:
            return Response(
                {"error": "Parâmetro 'id' é obrigatório."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Tenta buscar o usuário com base no ID fornecido
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "Usuário não encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Tenta buscar o funcionário associado ao usuário
        funcionario = None
        try:
            funcionario = user.funcionario
        except Funcionario.DoesNotExist:
            return Response(
                {"error": "Funcionário não encontrado para este usuário."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Monta o retorno com as informações do User, Pessoa, e Funcionario
        response_data = {
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "restaurante_id": user.restaurante.id if user.restaurante else None,
            },
            "pessoa": {
                "nome": funcionario.nome,
                "telefone": funcionario.telefone,
                "email": funcionario.email,
                "cpf_cnpj": funcionario.cpf_cnpj,
                # Adicione outros campos que você deseja retornar de Pessoa
            },
            "funcionario": {
                "pode_acessar_pdv": funcionario.pode_acessar_pdv,
                "pode_acessar_garcom": funcionario.pode_acessar_garcom,
                "pode_acessar_cozinha": funcionario.pode_acessar_cozinha,
                # Adicione outros campos que você deseja retornar de Funcionario
            },
        }

        return Response(response_data, status=status.HTTP_200_OK)


class RestauranteLoginView(APIView):
    def post(self, request):
        # Get email and password from the request
        email = request.data.get("email")
        senha = request.data.get("password")

        # Check if email and password are provided
        if not email or not senha:
            return Response(
                {"error": "E-mail e senha são obrigatórios."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Authenticate the user
        user = authenticate(request, username=email, password=senha)
        if user is not None:
            # Check if the user is linked to an employee
            try:
                funcionario = user.funcionario
            except Funcionario.DoesNotExist:
                return Response(
                    {"error": "Usuário não está vinculado a nenhum funcionário."},
                    status=status.HTTP_401_UNAUTHORIZED,
                )

            # Generate JWT token
            refresh = RefreshToken.for_user(user)
            return Response(
                {
                    "refresh": str(refresh),
                    "access": str(refresh.access_token),
                    "user_id": user.id,
                    "restaurante_id": funcionario.restaurante.id,
                    "roles": {
                        "pode_acessar_pdv": funcionario.pode_acessar_pdv,
                        "pode_acessar_garcom": funcionario.pode_acessar_garcom,
                        "pode_acessar_cozinha": funcionario.pode_acessar_cozinha,
                    },
                },
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"error": "E-mail ou senha incorretos."},
                status=status.HTTP_401_UNAUTHORIZED,
            )


class RestauranteViewSet(viewsets.ModelViewSet):
    queryset = Restaurante.objects.all()
    serializer_class = RestauranteSerializer


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer


class MesaViewSet(viewsets.ModelViewSet):
    queryset = Mesa.objects.all()
    serializer_class = MesaSerializer

    @action(detail=True, methods=["get"])
    def details(self, request, pk=None):
        mesa = self.get_object()
        serializer = MesaEPedidosSerializer(mesa)
        return Response(serializer.data)


class PratoViewSet(viewsets.ModelViewSet):
    serializer_class = PratoSerializer

    def get_queryset(self):
        restaurante_id = self.kwargs.get(
            "restaurante_id"
        )  # Obtém o ID do restaurante da URL
        return Prato.objects.filter(restaurante=restaurante_id, mostrar_cardapio=True)


class PessoaViewSet(viewsets.ModelViewSet):
    queryset = Pessoa.objects.all()
    serializer_class = PessoaSerializer


class NotificacaoViewSet(viewsets.ModelViewSet):
    serializer_class = NotificacaoSerializer

    def get_queryset(self):
        restaurante_id = self.kwargs.get("restaurante_id")
        return Notificacao.objects.filter(restaurante=restaurante_id)


class PratoPedidoViewSet(viewsets.ModelViewSet):
    queryset = PratoPedido.objects.all()
    serializer_class = PratoPedidoSerializer


class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer


class PratoSimpleViewSet(viewsets.ModelViewSet):
    queryset = Prato.objects.all()
    serializer_class = PratoSimpleSerializer


class CaixaViewSet(viewsets.ModelViewSet):
    queryset = Caixa.objects.all()
    serializer_class = CaixaSerializer


class DespesaViewSet(viewsets.ModelViewSet):
    queryset = Despesa.objects.all()
    serializer_class = DespesaSerializer


class MetodoPagamentoViewSet(viewsets.ModelViewSet):
    queryset = MetodoPagamento.objects.all()
    serializer_class = MetodoPagamentoSerializer


class DeliveryViewSet(viewsets.ModelViewSet):
    queryset = Delivery.objects.all()
    serializer_class = DeliverySerializer


class AcrescimoViewSet(viewsets.ModelViewSet):
    queryset = Acrescimo.objects.all()
    serializer_class = AcrescimoSerializer
